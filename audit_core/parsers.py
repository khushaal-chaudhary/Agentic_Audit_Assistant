from __future__ import annotations

import csv
import hashlib
import re
import unicodedata
import xml.etree.ElementTree as ET
from dataclasses import dataclass, replace
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from functools import lru_cache
from pathlib import Path
from typing import Iterable, Iterator

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

from .models import EvidenceRef


@lru_cache(maxsize=256)
def file_sha256(path: str) -> str:
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()


def relative_name(path: Path, root: Path) -> str:
    return path.relative_to(root).as_posix()


def decode_text(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in _text_encodings():
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _text_encodings() -> tuple[str, ...]:
    return ("utf-8-sig", "cp1252", "utf-8")


def detect_text_encoding(path: Path) -> str:
    with path.open("rb") as stream:
        sample = stream.read(64 * 1024)
    for encoding in _text_encodings():
        try:
            sample.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    return "utf-8"


def normalize_text(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return "".join(char for char in text if not unicodedata.combining(char)).casefold().strip()


def parse_decimal(value: object) -> Decimal:
    text = str(value or "").strip().replace("\u00a0", "").replace("€", "")
    if not text:
        return Decimal("0")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Cannot parse decimal value: {value!r}") from exc


def parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for pattern in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


@dataclass(frozen=True)
class SourceRow:
    source: Path
    root: Path
    row_number: int
    data: dict[str, str]
    raw: str
    sheet: str | None = None
    cells: dict[str, str] | None = None

    def evidence(self, *columns: str, query: str | None = None) -> EvidenceRef:
        excerpt = "; ".join(f"{column}={self.data.get(column, '')}" for column in columns)
        return EvidenceRef(
            document=relative_name(self.source, self.root),
            locator_type="cell" if self.sheet else "row",
            row=self.row_number,
            columns=list(columns),
            sheet=self.sheet,
            cell_range=self._cell_range(columns),
            excerpt=excerpt or self.raw[:500],
            query=query,
            sha256=file_sha256(str(self.source)),
        )

    def _cell_range(self, columns: Iterable[str]) -> str | None:
        if not self.cells:
            return None
        selected = [self.cells[column] for column in columns if column in self.cells]
        return ",".join(selected) if selected else None


@dataclass(frozen=True)
class PdfExtraction:
    page_count: int
    extracted_pages: int
    passages: list[EvidenceRef]
    error: str | None = None

    @property
    def status(self) -> str:
        if self.extracted_pages == 0:
            return "unreadable"
        if self.extracted_pages < self.page_count:
            return "partial"
        return "native"


def _pdf_page_passages(
    path: Path,
    root: Path,
    page_number: int,
    text: str,
    *,
    max_chars: int = 1200,
) -> list[EvidenceRef]:
    lines = [(index, line.rstrip()) for index, line in enumerate(text.splitlines(), start=1)]
    non_empty = [(index, line) for index, line in lines if line.strip()]
    passages: list[EvidenceRef] = []
    block: list[tuple[int, str]] = []
    size = 0

    def append_block() -> None:
        if not block:
            return
        start, end = block[0][0], block[-1][0]
        passages.append(
            EvidenceRef(
                document=relative_name(path, root),
                locator_type="page",
                page=page_number,
                passage=f"lines:{start}-{end}",
                excerpt="\n".join(line for _, line in block),
                sha256=file_sha256(str(path)),
            )
        )

    for line_number, line in non_empty:
        added = len(line) + (1 if block else 0)
        if block and size + added > max_chars:
            append_block()
            block = []
            size = 0
        block.append((line_number, line))
        size += added
    append_block()
    return passages


def read_pdf_passages(path: Path, root: Path) -> PdfExtraction:
    """Extract native PDF text into exact page-and-line passages.

    OCR is deliberately out of scope. Image-only or otherwise unreadable pages remain visible in
    coverage as unreadable instead of being silently treated as extracted.
    """
    try:
        reader = PdfReader(path, strict=False)
        if reader.is_encrypted and not reader.decrypt(""):
            return PdfExtraction(0, 0, [], "Encrypted PDF cannot be opened without a password")
    except Exception as exc:
        return PdfExtraction(0, 0, [], f"PDF could not be opened: {exc}")

    passages: list[EvidenceRef] = []
    extracted_pages = 0
    page_errors: list[int] = []
    for page_number, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception:
            page_errors.append(page_number)
            continue
        if not text.strip():
            continue
        extracted_pages += 1
        passages.extend(_pdf_page_passages(path, root, page_number, text))
    error = (
        "Native text extraction failed on page(s): " + ", ".join(map(str, page_errors))
        if page_errors
        else None
    )
    return PdfExtraction(len(reader.pages), extracted_pages, passages, error)


def gdpdu_headers(folder: Path, filename: str) -> list[str]:
    tree = ET.parse(folder / "index.xml")
    for table in tree.iter():
        if table.tag.split("}")[-1] != "Table":
            continue
        url = next(
            (child.text for child in table if child.tag.split("}")[-1] == "URL"), None
        )
        if url != filename:
            continue
        return [
            child.text or ""
            for column in table.iter()
            if column.tag.split("}")[-1] == "VariableColumn"
            for child in column
            if child.tag.split("}")[-1] == "Name"
        ]
    raise ValueError(f"No GDPdU schema for {filename}")


def iter_semicolon(
    path: Path,
    root: Path,
    headers: list[str] | None = None,
) -> Iterator[SourceRow]:
    """Stream a semicolon table while preserving one-based physical row locators."""
    encoding = detect_text_encoding(path)
    with path.open("r", encoding=encoding, errors="replace", newline="") as stream:
        reader = csv.reader(stream, delimiter=";", quotechar='"')
        if headers is None:
            try:
                fieldnames = [value.strip() for value in next(reader)]
            except StopIteration:
                return
            first_row_number = 2
        else:
            fieldnames = headers
            first_row_number = 1
        for row_number, values in enumerate(reader, start=first_row_number):
            padded = values + [""] * max(0, len(fieldnames) - len(values))
            data = dict(zip(fieldnames, padded, strict=False))
            yield SourceRow(
                path,
                root,
                row_number,
                data,
                ";".join(values),
            )


def read_semicolon(path: Path, root: Path, headers: list[str] | None = None) -> list[SourceRow]:
    return list(iter_semicolon(path, root, headers))


def read_xlsx_table(path: Path, root: Path, required_header: str) -> list[SourceRow]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    rows: list[SourceRow] = []
    for sheet in workbook.worksheets:
        header_index = None
        headers: list[str] = []
        for row in sheet.iter_rows():
            values = [str(cell.value or "").strip() for cell in row]
            if required_header in values:
                header_index = row[0].row
                headers = values
                break
        if header_index is None:
            continue
        for row in sheet.iter_rows(min_row=header_index + 1):
            values = [str(cell.value or "").strip() for cell in row]
            if not any(values):
                continue
            data = dict(zip(headers, values, strict=False))
            cells = {header: cell.coordinate for header, cell in zip(headers, row, strict=False)}
            rows.append(
                SourceRow(
                    path,
                    root,
                    row[0].row,
                    data,
                    " | ".join(values),
                    sheet.title,
                    cells,
                )
            )
    return rows


def canonicalize_rows(
    rows: list[SourceRow],
    header_map: dict[str, str],
) -> list[SourceRow]:
    canonical: list[SourceRow] = []
    for row in rows:
        data = dict(row.data)
        cells = dict(row.cells) if row.cells else None
        for canonical_name, actual_name in header_map.items():
            data[canonical_name] = row.data.get(actual_name, "")
            if cells is not None and actual_name in cells:
                cells[canonical_name] = cells[actual_name]
        canonical.append(replace(row, data=data, cells=cells))
    return canonical


def read_docx_passages(path: Path, root: Path) -> list[EvidenceRef]:
    document = Document(path)
    refs: list[EvidenceRef] = []
    for index, paragraph in enumerate(document.paragraphs, start=1):
        text = paragraph.text.strip()
        if not text:
            continue
        refs.append(
            EvidenceRef(
                document=relative_name(path, root),
                locator_type="passage",
                passage=f"paragraph:{index}",
                excerpt=text,
                sha256=file_sha256(str(path)),
            )
        )
    return refs


def locate_dossier_root(path: Path) -> Path:
    candidate = path.resolve()
    if (candidate / "Sachkonten" / "index.xml").exists():
        return candidate
    roots: set[Path] = set()
    signature = {"SACHKONTONUMMER", "BUCHUNGSNUMMER", "BUCHUNGSBETRAG"}
    for index_path in candidate.rglob("index.xml"):
        try:
            tree = ET.parse(index_path)
        except (ET.ParseError, OSError):
            continue
        headers = {
            child.text or ""
            for column in tree.iter()
            if column.tag.split("}")[-1] == "VariableColumn"
            for child in column
            if child.tag.split("}")[-1] == "Name"
        }
        if signature <= headers:
            roots.add(index_path.parent.parent)
    if len(roots) != 1:
        raise ValueError(f"Expected one GDPdU dossier below {path}, found {len(roots)}")
    return roots.pop()


def extract_payment_threshold(passages: list[EvidenceRef]) -> tuple[Decimal, EvidenceRef] | None:
    for ref in passages:
        normalized = normalize_text(ref.excerpt)
        if "zahlungsfreig" not in normalized and "vier-augen" not in normalized:
            continue
        match = re.search(r"(?:ab|uber)\s+([0-9][0-9.\s]*)\s*eur", normalized)
        if match:
            control_number = match.group(1).replace(" ", "").replace(".", "")
            return parse_decimal(control_number), ref
    return None


def extract_jet_threshold(passages: list[EvidenceRef]) -> tuple[Decimal, EvidenceRef] | None:
    for ref in passages:
        normalized = normalize_text(ref.excerpt)
        match = re.search(
            r"(?:nichtaufgriffsgrenze|clearly trivial threshold)(?:\s+jet)?\s*:?\s*"
            r"([0-9][0-9.,\s]*)\s*eur",
            normalized,
        )
        if match:
            control_number = re.sub(r"[^0-9]", "", match.group(1))
            return parse_decimal(control_number), ref
    return None
